"""
master.xlsx に signal-sensor / feedback-collector が使う 2シートを追加。
- candidates: 外部信号から AIエージェントが提案する候補ナレッジ（status=candidate）
- effectiveness: 各エントリの TP/FP/skipped 集計（評価のたびに更新）
"""
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_ROOT = os.environ.get("CLAUDE_PLUGIN_ROOT") or os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
XLSX = os.path.join(_ROOT, "knowledge", "master.xlsx")
NAVY = "0E2A47"; WHITE = "FFFFFF"; BORDER = "D9E2EC"

wb = load_workbook(XLSX)
print(f"既存シート: {wb.sheetnames}")

# ---- candidates シート ----
if "candidates" not in wb.sheetnames:
    ws = wb.create_sheet("candidates")
    headers = [
        "candidate_id", "detected_at", "source_type", "source_url",
        "raw_summary", "proposed_id", "proposed_title",
        "proposed_phase", "proposed_gate", "proposed_severity",
        "tech_relevance_score", "rationale",
        "status",  # candidate / promoted / rejected
        "reviewed_by", "reviewed_at", "decision_note",
    ]
    ws.append(headers)
    header_font = Font(name="Yu Gothic UI", size=11, bold=True, color=WHITE)
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    thin = Side(border_style="thin", color=BORDER)
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for i in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=i)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    widths = [16, 17, 14, 30, 50, 22, 40, 10, 22, 10, 10, 50, 12, 14, 17, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    print("✅ candidates シート作成")

# ---- effectiveness シート ----
if "effectiveness" not in wb.sheetnames:
    ws = wb.create_sheet("effectiveness")
    headers = [
        "id", "true_positive", "false_positive", "true_negative",
        "skipped", "last_finding", "last_evaluated_project",
        "fp_rate_30d", "tp_rate_30d", "review_status",
    ]
    ws.append(headers)
    header_font = Font(name="Yu Gothic UI", size=11, bold=True, color=WHITE)
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    thin = Side(border_style="thin", color=BORDER)
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for i in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=i)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    widths = [28, 14, 14, 14, 10, 17, 30, 12, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    print("✅ effectiveness シート作成")

# ---- senses シート（取込履歴） ----
if "senses" not in wb.sheetnames:
    ws = wb.create_sheet("senses")
    headers = [
        "sensed_at", "trigger", "source_type", "source_url",
        "raw_title", "raw_severity", "tech_match", "decision",
        "linked_candidate_id",
    ]
    ws.append(headers)
    header_font = Font(name="Yu Gothic UI", size=11, bold=True, color=WHITE)
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    thin = Side(border_style="thin", color=BORDER)
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for i in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=i)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    widths = [17, 16, 14, 50, 50, 12, 14, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    print("✅ senses シート作成")

wb.save(XLSX)
print(f"\n保存: {XLSX}")
print(f"シート構成: {wb.sheetnames}")
