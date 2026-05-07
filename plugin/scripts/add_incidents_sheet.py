"""
master.xlsx に incidents シートを追加（Reflective Curator 用）。
自社で発生したインシデント・FP記録・気づきを構造化して蓄積する。
"""
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_ROOT = os.environ.get("CLAUDE_PLUGIN_ROOT") or os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
XLSX = os.path.join(_ROOT, "knowledge", "master.xlsx")
NAVY = "0E2A47"; WHITE = "FFFFFF"; BORDER = "D9E2EC"

wb = load_workbook(XLSX)
print(f"既存シート: {wb.sheetnames}")

if "incidents" not in wb.sheetnames:
    ws = wb.create_sheet("incidents")
    headers = [
        "incident_id",       # INC-yyyymmdd-NNN
        "recorded_at",
        "incident_type",     # incident / false-positive / near-miss / observation
        "project_name",
        "severity",          # critical / high / medium / low
        "summary",           # 1行要約
        "what_happened",     # 何が起きたか
        "root_cause",        # 根本原因
        "abstracted_lesson", # 抽象化された教訓
        "related_knowledge_ids",  # 関連する既存ナレッジID（カンマ区切り）
        "related_evidence",  # ファイル/PR/Issue/ログのリンク
        "recorded_by",
        "status",            # raw / processed / archived
        "processed_at",      # reflective-curator が処理した日時
        "candidate_id",      # 派生した candidate のID（あれば）
    ]
    ws.append(headers)
    header_font = Font(name="Yu Gothic UI", size=11, bold=True, color=WHITE)
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    thin = Side(border_style="thin", color=BORDER)
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for i in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=i)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    widths = [16, 17, 15, 20, 10, 50, 60, 60, 60, 30, 40, 14, 11, 17, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    print("✅ incidents シート作成")

if "reflections" not in wb.sheetnames:
    ws = wb.create_sheet("reflections")
    headers = [
        "reflection_id",     # REF-yyyymmdd-NNN
        "executed_at",
        "trigger",           # manual / scheduled / fp-rate-threshold
        "incidents_count",   # 対象 incidents 件数
        "fp_high_count",     # FP率 > 30% のエントリ数
        "candidates_emitted", # 派生した candidates 件数
        "summary",           # 内省結果のサマリ
        "report_path",       # 生成された Markdown レポートのパス
    ]
    ws.append(headers)
    header_font = Font(name="Yu Gothic UI", size=11, bold=True, color=WHITE)
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    thin = Side(border_style="thin", color=BORDER)
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for i in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=i)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    widths = [16, 17, 18, 12, 12, 16, 60, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    print("✅ reflections シート作成")

wb.save(XLSX)
print(f"\n保存: {XLSX}")
print(f"シート構成: {wb.sheetnames}")
