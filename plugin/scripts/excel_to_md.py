"""
master.xlsx → knowledge/{phases,cross-cutting}/*.md 一括再生成
- Checklist シートの status=active 行のみ Markdown 化
- candidates シートは Markdown 化しない（評価対象外）
- 既存の knowledge/*.md は削除してから再生成
"""
import os
import sys
import yaml
import shutil
from openpyxl import load_workbook

# プラグイン配置先（Claude Code Plugin として動かす場合は CLAUDE_PLUGIN_ROOT）
# 直接スクリプト実行時は、本スクリプトの 1 階層上をプラグイン ROOT とみなす
ROOT = os.environ.get("CLAUDE_PLUGIN_ROOT") or os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
XLSX = os.path.join(ROOT, "knowledge/master.xlsx")
KB_DIR = os.path.join(ROOT, "knowledge")

# 列順（v4.xlsxと同じ）
COL = {
    "No": 1, "ID": 2, "Severity": 3, "重要度": 4,
    "Phase": 5, "Gate": 6, "観点・確認内容": 7, "適用条件": 8,
    "OK基準": 9, "NG基準": 10, "確認方法・ツール": 11, "必要証跡": 12,
    "判定者": 13, "Status": 14,
    "Last Validated": 15, "Review Frequency": 16,
    "判定結果": 17, "コメント": 18,
}

def search_hints(gate):
    domain = gate.split("/")[0]
    hints = {
        "legal": "法務関連文書（典型: docs/legal/, docs/compliance/, contracts/, 社内文書管理リンク）",
        "compliance": "規格対応文書（典型: docs/compliance/{規格名}/, docs/audit/, セキュリティレビュー記録）",
        "cost": "コスト試算（典型: docs/cost/, docs/finops/, スプレッドシート, README内予算記載）",
        "strategy": "プロジェクト戦略文書（典型: docs/project/, README, kickoff資料）",
        "architecture": "アーキ設計（典型: docs/adr/, docs/architecture/, design/, 図ファイル）",
        "data": "データ設計（典型: docs/design/data/, docs/data/, schemas/, ERD）",
        "api": "API仕様（典型: docs/api/, openapi.yaml, swagger.json, GraphQL schema）",
        "messaging": "メッセージング設計（典型: docs/design/messaging/, docs/adr/）",
        "nfr": "非機能要件（典型: docs/nfr/, docs/sre/, SLO定義）",
        "env": "環境設計（典型: docs/infra/, terraform/, k8s/, env/）",
        "security": "セキュリティ設計・実装（典型: docs/security/, .github/dependabot.yml, src/auth/, src/middleware/）",
        "perf": "性能関連（典型: docs/performance/, benchmarks/, ソースコード）",
        "impl": "実装コード（src/, lib/, app/）",
        "ops": "運用関連（典型: docs/ops/, docs/runbook/, ops/, .github/workflows/）",
        "quality": "コード品質（typically: ESLint設定, .editorconfig, ソースコード）",
        "git": "Git運用（.gitignore, .github/, リポジトリ設定）",
        "test": "テスト関連（典型: tests/, __tests__/, docs/test/, CI設定）",
        "stage": "ステージング設定（典型: docs/stage/, env/staging/, 演習レポート）",
        "release": "リリース手順（典型: docs/release/, RELEASE.md, .github/workflows/release.yml）",
        "llm": "LLM運用（典型: docs/llm/, evals/, prompts/, llm-config.yaml）",
        "governance": "統制関連（典型: docs/governance/, docs/policy/, 社内ポータル）",
        "ux": "UX設計（典型: docs/ux/, design-system/, src/components/）",
        "doc": "ドキュメント整備状況（典型: docs/, README.md, .github/CODEOWNERS）",
    }
    return hints.get(domain, "プロジェクト全体のドキュメント・コードを横断的に探索")


def main():
    if not os.path.exists(XLSX):
        print(f"❌ master.xlsx not found: {XLSX}")
        sys.exit(1)

    wb = load_workbook(XLSX)
    if "Checklist" not in wb.sheetnames:
        print(f"❌ Checklist シートが master.xlsx に見つからない")
        sys.exit(1)

    ws = wb["Checklist"]

    # 既存 *.md を削除（INDEX.md と schema.yaml は保持）
    for sub in ["phases", "cross-cutting"]:
        d = os.path.join(KB_DIR, sub)
        if os.path.isdir(d):
            shutil.rmtree(d)
            os.makedirs(d, exist_ok=True)

    # Checklist シートを読込み
    count = 0
    skipped = 0
    phase_index = {p: [] for p in ["P0","P1","P2","P3","P4","P5","P6","cross-cutting"]}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        eid = row[COL["ID"]-1]
        if not eid:
            continue
        sev = row[COL["Severity"]-1]
        priority = row[COL["重要度"]-1]
        ph = row[COL["Phase"]-1]
        gate = row[COL["Gate"]-1]
        observation = row[COL["観点・確認内容"]-1]
        applies = row[COL["適用条件"]-1]
        ok = row[COL["OK基準"]-1]
        ng = row[COL["NG基準"]-1]
        tools = row[COL["確認方法・ツール"]-1]
        ev = row[COL["必要証跡"]-1]
        judge = row[COL["判定者"]-1]
        status = row[COL["Status"]-1]
        last_val = row[COL["Last Validated"]-1]
        review_freq = row[COL["Review Frequency"]-1]

        # active 以外はスキップ
        if status != "active":
            skipped += 1
            continue

        fm = {
            "id": eid,
            "title": "(取得失敗時はExcelを参照)",  # title はExcelに無いので observation から取得
            "phase": [ph] if ph != "X" else ["cross-cutting"],
            "gate": gate,
            "severity": sev,
            "priority": priority,
            "judge": judge,
            "status": status,
            "last_validated": str(last_val) if last_val else "2026-04-30",
            "review_frequency": review_freq or "quarterly",
            "applies_when": applies,
        }

        # IDからtitle近似（IDから推測 or observationから先頭抜粋）
        title_short = (observation or eid).split("\n")[0][:60]
        fm["title"] = title_short

        fm_yaml = yaml.dump(fm, allow_unicode=True, sort_keys=False, default_flow_style=False)
        domain = (gate or "").split("/")[0]
        hint = search_hints(gate or "")

        body = f"""# {eid}: {title_short}

## 観点・確認内容

{observation or ''}

## 適用条件

{applies or ''}

## OK基準

{ok or ''}

## NG基準

{ng or ''}

## 必要証跡

{ev or ''}

## AI自律確認の手順

**AI agentが探すべき場所のヒント（決め打ちではない）**:
{hint}

**AI が実行する確認ステップ**:
1. プロジェクトルートから上記ヒントに沿って `Glob`/`Grep` で関連ファイルを探索
2. ファイル内容を `Read` で確認、必須項目の網羅性を判定
3. 承認記録（PR Approve / 電子サイン / チケットID）の紐付けを確認
4. 上記「OK基準」を満たすか判定し、不足があれば「NG基準」と「不足要素」を引用付きで報告
5. 判定結果は `cited_knowledge: [{eid}]` を必ず付与

**確認方法・ツール**: {tools or ''}

## Humanレビュー観点

判定者が `{judge}` の場合の人間関与:
- AI のみ: サンプリング監査（誤検知率が10%超なら全件人間レビューに移行）
- Both: AI判定結果のレビュー＋最終承認

## 陳腐化判定基準

- 関連する規格・法令・主要ライブラリの改訂
- 自社で類似のインシデント発生時
- AI判定の False Positive 率 > 30% が3ヶ月続いた場合
- 上記いずれかが発生したら revalidate モードで再検証

## 関連ナレッジ

（同じGate '{gate}' のエントリ）
"""

        # 出力先
        if ph != "X":
            target_dir = os.path.join(KB_DIR, "phases", ph)
        else:
            target_dir = os.path.join(KB_DIR, "cross-cutting", domain)
        os.makedirs(target_dir, exist_ok=True)
        path = os.path.join(target_dir, f"{eid}.md")
        with open(path, "w", encoding="utf-8") as f:
            f.write("---\n")
            f.write(fm_yaml)
            f.write("---\n\n")
            f.write(body)

        bucket = "cross-cutting" if ph == "X" else ph
        phase_index[bucket].append((eid, title_short, sev, gate))
        count += 1

    # INDEX.md 再生成
    with open(os.path.join(KB_DIR, "INDEX.md"), "w", encoding="utf-8") as f:
        f.write("# Agentic Quality Gate — Knowledge Index\n\n")
        f.write(f"全 {count} エントリ（status=active のみ）。skip {skipped} 件（draft/deprecated）。\n\n")
        f.write(f"**Master**: `knowledge/master.xlsx`（編集元）\n")
        f.write(f"**Generated**: このファイルと配下 *.md は master.xlsx から自動生成（`scripts/excel_to_md.py`）\n\n")
        sev_order = {"critical":0,"high":1,"medium":2,"low":3}
        for ph in ["P0","P1","P2","P3","P4","P5","P6","cross-cutting"]:
            items = phase_index[ph]
            if not items:
                continue
            title = {
                "P0": "P0 構想・要件定義",
                "P1": "P1 アーキテクチャ・データ設計",
                "P2": "P2 実装",
                "P3": "P3 テスト",
                "P4": "P4 ステージング・受入",
                "P5": "P5 リリース",
                "P6": "P6 運用・進化",
                "cross-cutting": "Cross-cutting",
            }[ph]
            f.write(f"## {title}（{len(items)}件）\n\n")
            f.write("| ID | Severity | Title | Gate |\n")
            f.write("|---|---|---|---|\n")
            items_sorted = sorted(items, key=lambda x: (sev_order.get(x[2], 9), x[0]))
            for eid, t, sev, gate in items_sorted:
                emoji = {"critical":"🔴","high":"🟠","medium":"🟡","low":"⚪"}.get(sev, "⚪")
                sub = "phases/" + ph if ph != "cross-cutting" else "cross-cutting/" + (gate or "").split("/")[0]
                f.write(f"| [{eid}]({sub}/{eid}.md) | {emoji} {sev} | {t} | `{gate}` |\n")
            f.write("\n")

    print(f"✅ Generated {count} markdown files (skipped {skipped} non-active rows)")
    print(f"   Output: {KB_DIR}")


if __name__ == "__main__":
    main()
