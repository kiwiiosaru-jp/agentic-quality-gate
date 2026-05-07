"""
master.xlsx に Claude Security findings 取込用シートを追加。
"""
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_ROOT = os.environ.get("CLAUDE_PLUGIN_ROOT") or os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
XLSX = os.path.join(_ROOT, "knowledge", "master.xlsx")
NAVY = "0E2A47"; WHITE = "FFFFFF"; BORDER = "D9E2EC"

wb = load_workbook(XLSX)
print(f"既存シート: {wb.sheetnames}")

if "claude_security_findings" not in wb.sheetnames:
    ws = wb.create_sheet("claude_security_findings")
    headers = [
        "imported_at",
        "claude_security_finding_id",  # Claude Security の内部ID
        "mapped_aqg_id",                # マッピング済みの本仕組みID
        "category",                     # CS のカテゴリ
        "severity",                     # critical / high / medium
        "verdict",                      # Fail / Conditional
        "repo",                         # GitHub リポジトリ
        "file",                         # ファイルパス
        "line_start",                   # 行番号開始
        "line_end",                     # 行番号終端
        "description",                  # CS の説明
        "patch_available",              # true/false
        "patch_link",                   # CS UI で patch を見るリンク
        "confidence",                   # CS の信頼度
        "status",                       # imported / merged / dismissed / patched
        "reviewer",                     # 人間レビュアー
        "reviewed_at",
        "decision_note",
    ]
    ws.append(headers)
    header_font = Font(name="Yu Gothic UI", size=11, bold=True, color=WHITE)
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    thin = Side(border_style="thin", color=BORDER)
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for i in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=i)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    widths = [17, 22, 28, 22, 11, 11, 35, 50, 10, 10, 50, 12, 30, 11, 11, 14, 17, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"
    print("✅ claude_security_findings シート作成")

wb.save(XLSX)
print(f"\n保存: {XLSX}")
print(f"シート構成: {wb.sheetnames}")
